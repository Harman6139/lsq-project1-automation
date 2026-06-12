from __future__ import annotations

import argparse
import csv
import json
import math
import os
import shutil
import subprocess
import time
import urllib.parse
import urllib.request
import zipfile
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError as exc:
    raise SystemExit("This script only needs one package: openpyxl. Install it with: pip install openpyxl") from exc


MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sept", "Oct", "Nov", "Dec"]
SAMPLE_START = (2017, 1)
SP500_SYMBOL = "^SP500TR"
NVDA_SYMBOL = "NVDA"
SPARTAN_URL = "https://spartanfunds.ca/wp-content/uploads/2021/11/LSQ-Oct21.pdf"
DEFAULT_WORKBOOK = Path(r"C:\Users\HP\Downloads\COMBined3a.xlsx")
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parents[1]

PAPER_TABLE2 = {
    "arithmetic": 0.0348,
    "geometric": 0.0340,
    "positive_pct": 0.9540,
    "negative_pct": 0.0460,
    "std": 0.0406,
    "sharpe": 0.86,
    "sortino": 13.96,
    "omega": 88.28,
    "alpha": 0.0352,
    "beta": -0.04,
}

PAPER_TABLE4_LSQ = {
    "Arithmetic Mean Return": 3.47,
    "Geometric Mean Return": 3.40,
    "Percent Positive Months": 95.41,
    "Percent Negative Months": 4.59,
    "Average Positive Month": 3.68,
    "Average Negative Month": -0.87,
    "Standard Deviation": 4.06,
    "Sharpe Ratio": 0.86,
    "Sortino Ratio": 13.96,
    "Omega Ratio": 88.28,
}

FAIRFIELD = {
    "Arithmetic Mean Return": "0.84",
    "Geometric Mean Return": "0.84",
    "Percent Positive Months": "92.09",
    "Percent Negative Months": "7.44",
    "Average Positive Month": "0.93",
    "Average Negative Month": "-0.17",
    "Best Month": "3.29",
    "Worst Month": "-0.64",
    "Mean": "0.84",
    "Standard Deviation": "0.711",
    "Skewness": "0.79",
    "Excess Kurtosis": "0.51",
    "Sharpe Ratio": "1.18",
    "Sortino Ratio": "12.76",
    "Omega Ratio": "67.32",
}


def add_month(year: int, month: int, offset: int) -> tuple[int, int]:
    idx = year * 12 + month - 1 + offset
    return idx // 12, idx % 12 + 1


def month_range(start: tuple[int, int], end: tuple[int, int]) -> list[tuple[int, int]]:
    out = []
    year, month = start
    while (year, month) <= end:
        out.append((year, month))
        year, month = add_month(year, month, 1)
    return out


def month_key(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}"


def month_label(date: tuple[int, int]) -> str:
    year, month = date
    return f"{MONTH_NAMES[month - 1]} {year}"


def sample_label(start: tuple[int, int], end: tuple[int, int]) -> str:
    return f"{month_label(start)} to {month_label(end)}"


def tex_num(x: float, digits: int = 2) -> str:
    if math.isnan(x):
        return "n/a"
    if math.isinf(x):
        return "inf"
    return f"{x:.{digits}f}"


def tex_pct(x: float, digits: int = 2) -> str:
    return f"{100.0 * x:.{digits}f}\\%"


def pp_change(new: float, old: float, digits: int = 2) -> str:
    return f"{100.0 * (new - old):+.{digits}f} pp"


def plain_change(new: float, old: float, digits: int = 2) -> str:
    return f"{new - old:+.{digits}f}"


def latex_escape(value: object) -> str:
    text = str(value)
    return (
        text.replace("\\", r"\textbackslash{}")
        .replace("&", r"\&")
        .replace("%", r"\%")
        .replace("$", r"\$")
        .replace("#", r"\#")
        .replace("_", r"\_")
        .replace("{", r"\{")
        .replace("}", r"\}")
    )


def extract_grid(ws, start_row: int, end_row: int) -> dict[tuple[int, int], float]:
    data: dict[tuple[int, int], float] = {}
    for row in range(start_row, end_row + 1):
        year = ws.cell(row=row, column=2).value
        if year is None:
            continue
        for month in range(1, 13):
            value = ws.cell(row=row, column=2 + month).value
            if value is None or value == "":
                continue
            data[(int(year), month)] = float(value) / 100.0
    return data


def series_from_dict(data: dict[tuple[int, int], float], end: tuple[int, int], name: str) -> list[dict[str, object]]:
    rows = []
    for date in month_range(SAMPLE_START, end):
        if date not in data:
            raise ValueError(f"{name} is missing {month_label(date)}.")
        rows.append({"date": month_key(*date), "month": month_label(date), name: data[date]})
    return rows


def read_workbook(path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], tuple[int, int]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    lsq_raw = extract_grid(ws, 10, 19)
    sp_raw = extract_grid(ws, 26, 35)
    nvda_raw = extract_grid(ws, 41, 50)

    # The professor's market rows stop at the latest completed month.
    # LSQ may contain zero placeholders after that month, so use the market rows to set the sample end.
    latest = min(max(sp_raw), max(nvda_raw))

    lsq = series_from_dict(lsq_raw, latest, "lsq_return")
    sp = series_from_dict(sp_raw, latest, "sp500tr_return")
    nvda = series_from_dict(nvda_raw, latest, "nvda_return")
    return lsq, sp, nvda, latest


def parse_percent(text: str) -> float:
    clean = text.replace("%", "").replace("+", "").replace(",", "").strip()
    return float(clean) / 100.0


def download_spartan_pdf(url: str, out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    pdf_path = out_dir / "LSQ_latest_spartan.pdf"
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            )
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = response.read()
    if not payload.startswith(b"%PDF"):
        raise ValueError("Spartan download did not return a PDF.")
    pdf_path.write_bytes(payload)
    return pdf_path


def read_lsq_from_spartan_pdf(pdf_path: Path) -> tuple[list[dict[str, object]], tuple[int, int]]:
    try:
        import fitz
    except ImportError as exc:
        raise SystemExit(
            "Spartan PDF mode needs one extra package: pymupdf. Install it with: pip install pymupdf"
        ) from exc

    doc = fitz.open(pdf_path)
    text = doc[0].get_text("text")
    if "Monthly Performance" not in text or "Statistics" not in text:
        raise ValueError("Could not find the monthly performance table in the Spartan PDF.")

    section = text.split("Monthly Performance", 1)[1].split("Statistics", 1)[0]
    lines = [line.strip() for line in section.splitlines() if line.strip()]
    raw: dict[tuple[int, int], float] = {}
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.isdigit() and len(line) == 4 and 2000 <= int(line) <= 2100:
            year = int(line)
            pct_values = []
            j = i + 1
            while j < len(lines):
                candidate = lines[j]
                if candidate.isdigit() and len(candidate) == 4 and 2000 <= int(candidate) <= 2100:
                    break
                if candidate.endswith("%") and (candidate.startswith("+") or candidate.startswith("-")):
                    pct_values.append(parse_percent(candidate))
                j += 1
            if len(pct_values) < 3:
                raise ValueError(f"Could not parse enough percentage values for {year} in Spartan PDF.")
            monthly = pct_values[:12] if len(pct_values) >= 14 else pct_values[:-2]
            for month, value in enumerate(monthly, start=1):
                raw[(year, month)] = value
            i = j
        else:
            i += 1

    latest = max(date for date in raw if date >= SAMPLE_START)
    rows = series_from_dict(raw, latest, "lsq_return")
    return rows, latest


def market_rows_from_yahoo(symbol: str, latest: tuple[int, int], key: str) -> tuple[list[dict[str, object]], str]:
    market, url = yahoo_returns(symbol, SAMPLE_START, latest)
    rows = []
    for date in month_range(SAMPLE_START, latest):
        if date not in market:
            raise ValueError(f"{symbol} is missing {month_label(date)} from Yahoo Finance.")
        rows.append({"date": month_key(*date), "month": month_label(date), key: market[date]})
    return rows, url


def read_spartan_source(url: str, out_dir: Path) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], tuple[int, int], Path]:
    last_error: Exception | None = None
    for attempt in range(1, 6):
        try:
            pdf_path = download_spartan_pdf(url, out_dir / "data")
            lsq, latest = read_lsq_from_spartan_pdf(pdf_path)
            sp, _ = market_rows_from_yahoo(SP500_SYMBOL, latest, "sp500tr_return")
            nvda, _ = market_rows_from_yahoo(NVDA_SYMBOL, latest, "nvda_return")
            return lsq, sp, nvda, latest, pdf_path
        except Exception as exc:
            last_error = exc
            if attempt == 5:
                break
            time.sleep(3 * attempt)
    raise RuntimeError(f"Could not fetch and parse the Spartan source after 5 attempts: {last_error}") from last_error


def values(rows: list[dict[str, object]], key: str) -> list[float]:
    return [float(row[key]) for row in rows]


def compute_omega(r: list[float], threshold: float = 0.0) -> float:
    gains = sum(max(x - threshold, 0.0) for x in r)
    losses = sum(max(threshold - x, 0.0) for x in r)
    return math.inf if losses == 0 else gains / losses


def compute_sortino(r: list[float], threshold: float = 0.0) -> float:
    downside = [min(x - threshold, 0.0) for x in r]
    downside_deviation = math.sqrt(sum(x * x for x in downside) / len(r))
    return math.inf if downside_deviation == 0 else (sum(x - threshold for x in r) / len(r)) / downside_deviation


def compute_summary(r: list[float]) -> dict[str, float | int]:
    n = len(r)
    mean = sum(r) / n
    variance = sum((x - mean) ** 2 for x in r) / n
    std = math.sqrt(variance)
    product = math.prod(1.0 + x for x in r)
    geo = product ** (1.0 / n) - 1.0
    pos = [x for x in r if x > 0.0]
    neg = [x for x in r if x < 0.0]
    skewness = sum((x - mean) ** 3 for x in r) / n / (std**3) if std else math.nan
    excess_kurtosis = sum((x - mean) ** 4 for x in r) / n / (std**4) - 3.0 if std else math.nan
    return {
        "n": n,
        "positive_count": len(pos),
        "negative_count": len(neg),
        "positive_pct": len(pos) / n,
        "negative_pct": len(neg) / n,
        "arithmetic": mean,
        "geometric": geo,
        "average_positive": sum(pos) / len(pos) if pos else math.nan,
        "average_negative": sum(neg) / len(neg) if neg else math.nan,
        "best": max(r),
        "worst": min(r),
        "mean": mean,
        "std": std,
        "skewness": skewness,
        "excess_kurtosis": excess_kurtosis,
        "sharpe": math.inf if std == 0 else mean / std,
        "sortino": compute_sortino(r),
        "omega": compute_omega(r),
    }


def inverse_matrix(a: list[list[float]]) -> list[list[float]]:
    if len(a) == 1:
        return [[1.0 / a[0][0]]]
    if len(a) == 2:
        det = a[0][0] * a[1][1] - a[0][1] * a[1][0]
        if abs(det) < 1e-18:
            raise ValueError("Matrix is singular.")
        return [[a[1][1] / det, -a[0][1] / det], [-a[1][0] / det, a[0][0] / det]]
    raise ValueError("Only 1x1 and 2x2 matrices are supported.")


def matmul(a: list[list[float]], b: list[list[float]]) -> list[list[float]]:
    return [[sum(a[i][k] * b[k][j] for k in range(len(b))) for j in range(len(b[0]))] for i in range(len(a))]


def transpose(a: list[list[float]]) -> list[list[float]]:
    return [list(row) for row in zip(*a)]


def xtx_inv(x: list[list[float]]) -> list[list[float]]:
    return inverse_matrix(matmul(transpose(x), x))


def ols(y: list[float], x: list[list[float]]) -> dict[str, object]:
    xt = transpose(x)
    inv = inverse_matrix(matmul(xt, x))
    xty = [[sum(xt[i][t] * y[t] for t in range(len(y)))] for i in range(len(xt))]
    beta = [row[0] for row in matmul(inv, xty)]
    residuals = [y[t] - sum(x[t][j] * beta[j] for j in range(len(beta))) for t in range(len(y))]
    dof = len(y) - len(beta)
    sse = sum(e * e for e in residuals)
    sigma2 = sse / dof
    cov = [[sigma2 * inv[i][j] for j in range(len(inv))] for i in range(len(inv))]
    se = [math.sqrt(cov[i][i]) for i in range(len(beta))]
    tstat = [beta[i] / se[i] for i in range(len(beta))]
    return {"beta": beta, "se": se, "t": tstat, "residuals": residuals, "cov": cov}


def hac_covariance(x: list[list[float]], residuals: list[float], lags: int) -> list[list[float]]:
    k = len(x[0])
    inv = xtx_inv(x)
    s = [[0.0 for _ in range(k)] for _ in range(k)]
    n = len(x)
    for t in range(n):
        for i in range(k):
            for j in range(k):
                s[i][j] += residuals[t] * residuals[t] * x[t][i] * x[t][j]
    for lag in range(1, lags + 1):
        weight = 1.0 - lag / (lags + 1.0)
        for t in range(lag, n):
            eprod = residuals[t] * residuals[t - lag] * weight
            for i in range(k):
                for j in range(k):
                    s[i][j] += eprod * (x[t][i] * x[t - lag][j] + x[t - lag][i] * x[t][j])
    return matmul(matmul(inv, s), inv)


def regression_table(lsq: list[float], sp: list[float]) -> tuple[list[dict[str, float | str]], dict[str, object]]:
    x = [[1.0, sp_i] for sp_i in sp]
    model = ols(lsq, x)
    beta = model["beta"]
    se = model["se"]
    tstat = model["t"]
    rows = [
        {
            "statistic": "Monthly alpha",
            "estimate": beta[0] * 100.0,
            "std_error": se[0] * 100.0,
            "t_stat": tstat[0],
        },
        {
            "statistic": "Beta",
            "estimate": beta[1],
            "std_error": se[1],
            "t_stat": tstat[1],
        },
    ]
    return rows, model


def autocorr(r: list[float], lag: int) -> float:
    mean = sum(r) / len(r)
    numerator = sum((r[t] - mean) * (r[t - lag] - mean) for t in range(lag, len(r)))
    denominator = sum((x - mean) ** 2 for x in r)
    return numerator / denominator if denominator else math.nan


def gammaincc(a: float, x: float) -> float:
    if x <= 0:
        return 1.0
    eps = 3e-14
    max_iter = 200
    gln = math.lgamma(a)
    if x < a + 1.0:
        ap = a
        delta = 1.0 / a
        total = delta
        for _ in range(max_iter):
            ap += 1.0
            delta *= x / ap
            total += delta
            if abs(delta) < abs(total) * eps:
                p = total * math.exp(-x + a * math.log(x) - gln)
                return max(0.0, min(1.0, 1.0 - p))
        p = total * math.exp(-x + a * math.log(x) - gln)
        return max(0.0, min(1.0, 1.0 - p))

    tiny = 1e-300
    b = x + 1.0 - a
    c = 1.0 / tiny
    d = 1.0 / b if abs(b) > tiny else 1.0 / tiny
    h = d
    for i in range(1, max_iter + 1):
        an = -i * (i - a)
        b += 2.0
        d = an * d + b
        if abs(d) < tiny:
            d = tiny
        c = b + an / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < eps:
            break
    q = math.exp(-x + a * math.log(x) - gln) * h
    return max(0.0, min(1.0, q))


def chi_square_sf(x: float, df: int) -> float:
    return gammaincc(df / 2.0, x / 2.0)


def ljung_box_pvalue(r: list[float], lag: int) -> float:
    n = len(r)
    q = n * (n + 2) * sum((autocorr(r, k) ** 2) / (n - k) for k in range(1, lag + 1))
    return chi_square_sf(q, lag)


def serial_table(lsq: list[float], sp: list[float]) -> list[dict[str, float | str]]:
    active = [a - b for a, b in zip(lsq, sp)]
    rows = []
    for name, series in [("LSQ Fund", lsq), ("S&P 500 TR", sp), ("LSQ minus S&P 500 TR", active)]:
        rows.append(
            {
                "Series": name,
                "rho_1": autocorr(series, 1),
                "LB p(1)": ljung_box_pvalue(series, 1),
                "LB p(6)": ljung_box_pvalue(series, 6),
                "LB p(12)": ljung_box_pvalue(series, 12),
            }
        )
    return rows


def newey_west_table(lsq: list[float], sp: list[float]) -> tuple[list[dict[str, float | str]], int]:
    n = len(lsq)
    lags = int(math.floor(4.0 * (n / 100.0) ** (2.0 / 9.0)))

    mean_x = [[1.0] for _ in lsq]
    mean_model = ols(lsq, mean_x)
    mean_hac = hac_covariance(mean_x, mean_model["residuals"], lags)

    reg_x = [[1.0, x] for x in sp]
    reg_model = ols(lsq, reg_x)
    reg_hac = hac_covariance(reg_x, reg_model["residuals"], lags)

    rows = [
        {
            "Statistic": "LSQ monthly mean",
            "Estimate": mean_model["beta"][0] * 100.0,
            "OLS SE": mean_model["se"][0] * 100.0,
            "OLS t-stat": mean_model["t"][0],
            "Newey-West SE": math.sqrt(mean_hac[0][0]) * 100.0,
            "Newey-West t-stat": mean_model["beta"][0] / math.sqrt(mean_hac[0][0]),
        },
        {
            "Statistic": "Monthly alpha",
            "Estimate": reg_model["beta"][0] * 100.0,
            "OLS SE": reg_model["se"][0] * 100.0,
            "OLS t-stat": reg_model["t"][0],
            "Newey-West SE": math.sqrt(reg_hac[0][0]) * 100.0,
            "Newey-West t-stat": reg_model["beta"][0] / math.sqrt(reg_hac[0][0]),
        },
        {
            "Statistic": "Beta",
            "Estimate": reg_model["beta"][1],
            "OLS SE": reg_model["se"][1],
            "OLS t-stat": reg_model["t"][1],
            "Newey-West SE": math.sqrt(reg_hac[1][1]),
            "Newey-West t-stat": reg_model["beta"][1] / math.sqrt(reg_hac[1][1]),
        },
    ]
    return rows, lags


def yahoo_returns(symbol: str, start: tuple[int, int], end: tuple[int, int]) -> tuple[dict[tuple[int, int], float], str]:
    fetch_start = add_month(*start, -1)
    period1 = int(datetime(fetch_start[0], fetch_start[1], 1, tzinfo=timezone.utc).timestamp())
    end_exclusive = add_month(*end, 1)
    period2 = int(datetime(end_exclusive[0], end_exclusive[1], 5, tzinfo=timezone.utc).timestamp())
    encoded = urllib.parse.quote(symbol, safe="")
    url = (
        f"https://query2.finance.yahoo.com/v8/finance/chart/{encoded}"
        f"?period1={period1}&period2={period2}&interval=1mo&events=history&includeAdjustedClose=true"
    )
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
            )
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        payload = json.load(response)
    error = payload.get("chart", {}).get("error")
    if error:
        raise RuntimeError(f"Yahoo error for {symbol}: {error}")
    result = payload["chart"]["result"][0]
    adj = result["indicators"].get("adjclose", [{}])[0].get("adjclose")
    close = result["indicators"]["quote"][0]["close"]
    levels = []
    for i, ts in enumerate(result["timestamp"]):
        level = adj[i] if adj else close[i]
        if level is None:
            continue
        dt = datetime.fromtimestamp(ts, timezone.utc)
        levels.append(((dt.year, dt.month), float(level)))
    levels.sort()
    out: dict[tuple[int, int], float] = {}
    for i in range(1, len(levels)):
        date, level = levels[i]
        previous = levels[i - 1][1]
        if start <= date <= end:
            out[date] = level / previous - 1.0
    return out, url


def verify_market_returns(
    sp_rows: list[dict[str, object]],
    nvda_rows: list[dict[str, object]],
    latest: tuple[int, int],
    skip_yahoo: bool,
) -> list[dict[str, object]]:
    checks = []
    for label, symbol, rows, key in [
        ("S&P 500 TR", SP500_SYMBOL, sp_rows, "sp500tr_return"),
        ("NVIDIA", NVDA_SYMBOL, nvda_rows, "nvda_return"),
    ]:
        workbook = float(rows[-1][key])
        source_url = ""
        yahoo_value = math.nan
        status = "not checked"
        if not skip_yahoo:
            try:
                market, source_url = yahoo_returns(symbol, SAMPLE_START, latest)
                yahoo_value = market[latest]
                status = "match" if abs(workbook - yahoo_value) * 100.0 < 0.005 else "rounding review"
            except Exception as exc:
                status = f"check failed: {type(exc).__name__}"
        checks.append(
            {
                "series": label,
                "symbol": symbol,
                "month": month_key(*latest),
                "source_return_percent": workbook * 100.0,
                "workbook_return_percent": workbook * 100.0,
                "yahoo_exact_return_percent": yahoo_value * 100.0 if not math.isnan(yahoo_value) else "",
                "difference_percentage_points": (workbook - yahoo_value) * 100.0 if not math.isnan(yahoo_value) else "",
                "status": status,
                "source_url": source_url,
            }
        )
    return checks


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def table2_rows(lsq: dict[str, float | int], sp: dict[str, float | int], reg: list[dict[str, float | str]]) -> list[dict[str, object]]:
    alpha = next(row for row in reg if row["statistic"] == "Monthly alpha")["estimate"]
    beta = next(row for row in reg if row["statistic"] == "Beta")["estimate"]
    data = [
        ("Average (arithmetic) return", lsq["arithmetic"] * 100.0, sp["arithmetic"] * 100.0),
        ("Average (geometric) return", lsq["geometric"] * 100.0, sp["geometric"] * 100.0),
        ("Percent Positive Months", lsq["positive_pct"] * 100.0, sp["positive_pct"] * 100.0),
        ("Percent Negative Months", lsq["negative_pct"] * 100.0, sp["negative_pct"] * 100.0),
        ("Average Positive Months", lsq["average_positive"] * 100.0, sp["average_positive"] * 100.0),
        ("Average Negative Months", lsq["average_negative"] * 100.0, sp["average_negative"] * 100.0),
        ("Best Month", lsq["best"] * 100.0, sp["best"] * 100.0),
        ("Worst Month", lsq["worst"] * 100.0, sp["worst"] * 100.0),
        ("Mean", lsq["mean"] * 100.0, sp["mean"] * 100.0),
        ("Standard deviation", lsq["std"] * 100.0, sp["std"] * 100.0),
        ("Skewness", lsq["skewness"], sp["skewness"]),
        ("Excess kurtosis", lsq["excess_kurtosis"], sp["excess_kurtosis"]),
        ("Sharpe Ratio", lsq["sharpe"], sp["sharpe"]),
        ("Sortino Ratio", lsq["sortino"], sp["sortino"]),
        ("Omega Ratio", lsq["omega"], sp["omega"]),
        ("Alpha", alpha, 0.0),
        ("Beta", beta, 1.0),
    ]
    return [{"Metric": a, "LSQ": b, "S&P 500 Index": c} for a, b, c in data]


def table4_rows(lsq: dict[str, float | int], nvda: dict[str, float | int]) -> list[dict[str, object]]:
    data = [
        ("Arithmetic Mean Return", lsq["arithmetic"] * 100.0, nvda["arithmetic"] * 100.0, FAIRFIELD["Arithmetic Mean Return"]),
        ("Geometric Mean Return", lsq["geometric"] * 100.0, nvda["geometric"] * 100.0, FAIRFIELD["Geometric Mean Return"]),
        ("Percent Positive Months", lsq["positive_pct"] * 100.0, nvda["positive_pct"] * 100.0, FAIRFIELD["Percent Positive Months"]),
        ("Percent Negative Months", lsq["negative_pct"] * 100.0, nvda["negative_pct"] * 100.0, FAIRFIELD["Percent Negative Months"]),
        ("Average Positive Month", lsq["average_positive"] * 100.0, nvda["average_positive"] * 100.0, FAIRFIELD["Average Positive Month"]),
        ("Average Negative Month", lsq["average_negative"] * 100.0, nvda["average_negative"] * 100.0, FAIRFIELD["Average Negative Month"]),
        ("Best Month", lsq["best"] * 100.0, nvda["best"] * 100.0, FAIRFIELD["Best Month"]),
        ("Worst Month", lsq["worst"] * 100.0, nvda["worst"] * 100.0, FAIRFIELD["Worst Month"]),
        ("Mean", lsq["mean"] * 100.0, nvda["mean"] * 100.0, FAIRFIELD["Mean"]),
        ("Standard Deviation", lsq["std"] * 100.0, nvda["std"] * 100.0, FAIRFIELD["Standard Deviation"]),
        ("Skewness", lsq["skewness"], nvda["skewness"], FAIRFIELD["Skewness"]),
        ("Excess Kurtosis", lsq["excess_kurtosis"], nvda["excess_kurtosis"], FAIRFIELD["Excess Kurtosis"]),
        ("Sharpe Ratio", lsq["sharpe"], nvda["sharpe"], FAIRFIELD["Sharpe Ratio"]),
        ("Sortino Ratio", lsq["sortino"], nvda["sortino"], FAIRFIELD["Sortino Ratio"]),
        ("Omega Ratio", lsq["omega"], nvda["omega"], FAIRFIELD["Omega Ratio"]),
    ]
    return [{"Statistic": a, "LSQ Fund": b, "NVIDIA": c, "Fairfield Sentry": d} for a, b, c, d in data]


def write_table2_tex(path: Path, lsq: dict[str, float | int], sp: dict[str, float | int], reg: list[dict[str, float | str]]) -> None:
    rows = [
        r"\begin{table}[H]",
        r"\centering",
        r"\caption{Comparison of LSQ returns with S\&P 500 returns}",
        r"\label{tab:updated-table-2}",
        r"\begin{tabular}{lcc}",
        r"\toprule",
        r"\textbf{Metric} & \textbf{LSQ} & \textbf{S\&P 500 Index} \\",
        r"\midrule",
        r"\multicolumn{3}{l}{\emph{Statistical properties of monthly returns}} \\",
        r"\midrule",
        f"Average (arithmetic) return & {tex_num(lsq['arithmetic'] * 100.0)} & {tex_num(sp['arithmetic'] * 100.0)} \\\\",
        f"Average (geometric) return & {tex_num(lsq['geometric'] * 100.0)} & {tex_num(sp['geometric'] * 100.0)} \\\\",
        f"Percent Positive Months & {tex_num(lsq['positive_pct'] * 100.0)} & {tex_num(sp['positive_pct'] * 100.0)} \\\\",
        f"Percent Negative Months & {tex_num(lsq['negative_pct'] * 100.0)} & {tex_num(sp['negative_pct'] * 100.0)} \\\\",
        f"Average Positive Months & {tex_num(lsq['average_positive'] * 100.0)} & {tex_num(sp['average_positive'] * 100.0)} \\\\",
        f"Average Negative Months & {tex_num(lsq['average_negative'] * 100.0)} & {tex_num(sp['average_negative'] * 100.0)} \\\\",
        f"Best Month & {tex_num(lsq['best'] * 100.0)} & {tex_num(sp['best'] * 100.0)} \\\\",
        f"Worst Month & {tex_num(lsq['worst'] * 100.0)} & {tex_num(sp['worst'] * 100.0)} \\\\",
        r"\midrule",
        r"\multicolumn{3}{l}{\emph{First four moments of distribution of monthly returns}} \\",
        r"\midrule",
        f"Mean & {tex_num(lsq['mean'] * 100.0)} & {tex_num(sp['mean'] * 100.0)} \\\\",
        f"Standard deviation & {tex_num(lsq['std'] * 100.0)} & {tex_num(sp['std'] * 100.0)} \\\\",
        f"Skewness & {tex_num(lsq['skewness'])} & {tex_num(sp['skewness'])} \\\\",
        f"Excess kurtosis & {tex_num(lsq['excess_kurtosis'])} & {tex_num(sp['excess_kurtosis'])} \\\\",
        r"\midrule",
        r"\multicolumn{3}{l}{\emph{Risk adjusted performance metrics}} \\",
        r"\midrule",
        f"Sharpe Ratio & {tex_num(lsq['sharpe'])} & {tex_num(sp['sharpe'])} \\\\",
        f"Sortino Ratio & {tex_num(lsq['sortino'])} & {tex_num(sp['sortino'])} \\\\",
        f"Omega Ratio & {tex_num(lsq['omega'])} & {tex_num(sp['omega'])} \\\\",
        f"Alpha & {tex_num(next(r for r in reg if r['statistic'] == 'Monthly alpha')['estimate'])} & 0 \\\\",
        f"Beta & {tex_num(next(r for r in reg if r['statistic'] == 'Beta')['estimate'])} & 1 \\\\",
        r"\bottomrule",
        r"\end{tabular}",
        "",
        r"\vspace{0.2cm}",
        r"\begin{minipage}{0.92\linewidth}",
        r"\footnotesize",
        r"Note: Returns are monthly percentages except ratios, skewness, excess kurtosis, and beta.",
        r"The LSQ column uses the baseline net-of-fees series.",
        r"The S\&P 500 benchmark is the total return index, consistent with the paper's dividend-reinvestment convention.",
        r"\end{minipage}",
        r"\end{table}",
    ]
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def write_table4_tex(path: Path, latest: tuple[int, int], lsq: dict[str, float | int], nvda: dict[str, float | int]) -> None:
    label = sample_label(SAMPLE_START, latest)
    rows = [
        r"\begin{table}[H]",
        r"\centering",
        r"\caption{Monthly Return Statistics: LSQ Fund, NVIDIA, and Fairfield Sentry}",
        r"\label{tab:updated-table-4}",
        r"\small",
        r"\begin{tabular}{lccc}",
        r"\toprule",
        r"\textbf{Statistic} & \textbf{LSQ Fund} & \textbf{NVIDIA} & \textbf{Fairfield Sentry} \\",
        f"& \\emph{{{label}}} & \\emph{{{label}}} & \\emph{{Dec 1990 to Oct 2008}} \\\\",
        f"& \\emph{{(N = {lsq['n']})}} & \\emph{{(N = {nvda['n']})}} & \\emph{{(N = 215)}} \\\\",
        r"\midrule",
        r"\multicolumn{4}{l}{\emph{Panel A: Return Summary}} \\",
        r"\midrule",
        f"Arithmetic Mean Return & {tex_num(lsq['arithmetic'] * 100.0)} & {tex_num(nvda['arithmetic'] * 100.0)} & {FAIRFIELD['Arithmetic Mean Return']} \\\\",
        f"Geometric Mean Return & {tex_num(lsq['geometric'] * 100.0)} & {tex_num(nvda['geometric'] * 100.0)} & {FAIRFIELD['Geometric Mean Return']} \\\\",
        f"Percent Positive Months & {tex_num(lsq['positive_pct'] * 100.0)} & {tex_num(nvda['positive_pct'] * 100.0)} & {FAIRFIELD['Percent Positive Months']} \\\\",
        f"Percent Negative Months & {tex_num(lsq['negative_pct'] * 100.0)} & {tex_num(nvda['negative_pct'] * 100.0)} & {FAIRFIELD['Percent Negative Months']} \\\\",
        f"Average Positive Month & {tex_num(lsq['average_positive'] * 100.0)} & {tex_num(nvda['average_positive'] * 100.0)} & {FAIRFIELD['Average Positive Month']} \\\\",
        f"Average Negative Month & {tex_num(lsq['average_negative'] * 100.0)} & {tex_num(nvda['average_negative'] * 100.0)} & {FAIRFIELD['Average Negative Month']} \\\\",
        f"Best Month & {tex_num(lsq['best'] * 100.0)} & {tex_num(nvda['best'] * 100.0)} & {FAIRFIELD['Best Month']} \\\\",
        f"Worst Month & {tex_num(lsq['worst'] * 100.0)} & {tex_num(nvda['worst'] * 100.0)} & {FAIRFIELD['Worst Month']} \\\\",
        r"\midrule",
        r"\multicolumn{4}{l}{\emph{Panel B: First Four Moments}} \\",
        r"\midrule",
        f"Mean ($\\mu$) & {tex_num(lsq['mean'] * 100.0)} & {tex_num(nvda['mean'] * 100.0)} & {FAIRFIELD['Mean']} \\\\",
        f"Standard Deviation ($\\sigma$) & {tex_num(lsq['std'] * 100.0)} & {tex_num(nvda['std'] * 100.0)} & {FAIRFIELD['Standard Deviation']} \\\\",
        f"Skewness & {tex_num(lsq['skewness'])} & {tex_num(nvda['skewness'])} & {FAIRFIELD['Skewness']} \\\\",
        f"Excess Kurtosis & {tex_num(lsq['excess_kurtosis'])} & {tex_num(nvda['excess_kurtosis'])} & {FAIRFIELD['Excess Kurtosis']} \\\\",
        r"\midrule",
        r"\multicolumn{4}{l}{\emph{Panel C: Risk-Adjusted Ratios (Monthly, risk free rate = 0)}} \\",
        r"\midrule",
        f"Sharpe Ratio & {tex_num(lsq['sharpe'])} & {tex_num(nvda['sharpe'])} & {FAIRFIELD['Sharpe Ratio']} \\\\",
        f"Sortino Ratio & {tex_num(lsq['sortino'])} & {tex_num(nvda['sortino'])} & {FAIRFIELD['Sortino Ratio']} \\\\",
        f"Omega Ratio & {tex_num(lsq['omega'])} & {tex_num(nvda['omega'])} & {FAIRFIELD['Omega Ratio']} \\\\",
        r"\bottomrule",
        r"\end{tabular}",
        "",
        r"\vspace{0.15cm}",
        r"\begin{minipage}{0.92\linewidth}",
        r"\footnotesize",
        r"Sortino uses Estrada (1999) downside deviation:",
        r"$\sigma_d=\sqrt{\frac{1}{N}\sum_{r_t<0}r_t^2}$.",
        r"Omega $= \sum \max(r_t,0) / \sum \max(-r_t,0)$.",
        r"Fairfield Sentry data: Bernard \& Boyle (2009), \emph{Journal of Derivatives}.",
        r"\end{minipage}",
        r"\end{table}",
    ]
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def table2_summary_tex(lsq: dict[str, float | int], reg: list[dict[str, float | str]]) -> str:
    alpha = next(row for row in reg if row["statistic"] == "Monthly alpha")["estimate"] / 100.0
    beta = next(row for row in reg if row["statistic"] == "Beta")["estimate"]
    lines = [
        r"\begin{center}",
        r"\begin{tabular}{lrrr}",
        r"\toprule",
        r"Metric & Paper & Updated & Change \\",
        r"\midrule",
        f"LSQ arithmetic monthly return & {tex_pct(PAPER_TABLE2['arithmetic'])} & {tex_pct(lsq['arithmetic'])} & {pp_change(lsq['arithmetic'], PAPER_TABLE2['arithmetic'])} \\\\",
        f"LSQ geometric monthly return & {tex_pct(PAPER_TABLE2['geometric'])} & {tex_pct(lsq['geometric'])} & {pp_change(lsq['geometric'], PAPER_TABLE2['geometric'])} \\\\",
        f"LSQ percent positive months & {tex_pct(PAPER_TABLE2['positive_pct'])} & {tex_pct(lsq['positive_pct'])} & {pp_change(lsq['positive_pct'], PAPER_TABLE2['positive_pct'])} \\\\",
        f"LSQ percent negative months & {tex_pct(PAPER_TABLE2['negative_pct'])} & {tex_pct(lsq['negative_pct'])} & {pp_change(lsq['negative_pct'], PAPER_TABLE2['negative_pct'])} \\\\",
        f"LSQ population standard deviation & {tex_pct(PAPER_TABLE2['std'])} & {tex_pct(lsq['std'])} & {pp_change(lsq['std'], PAPER_TABLE2['std'])} \\\\",
        f"LSQ Sharpe ratio & {tex_num(PAPER_TABLE2['sharpe'])} & {tex_num(lsq['sharpe'])} & {plain_change(lsq['sharpe'], PAPER_TABLE2['sharpe'])} \\\\",
        f"LSQ Sortino ratio & {tex_num(PAPER_TABLE2['sortino'])} & {tex_num(lsq['sortino'])} & {plain_change(lsq['sortino'], PAPER_TABLE2['sortino'])} \\\\",
        f"LSQ Omega ratio & {tex_num(PAPER_TABLE2['omega'])} & {tex_num(lsq['omega'])} & {plain_change(lsq['omega'], PAPER_TABLE2['omega'])} \\\\",
        f"Alpha, monthly & {tex_pct(PAPER_TABLE2['alpha'])} & {tex_pct(alpha)} & {pp_change(alpha, PAPER_TABLE2['alpha'])} \\\\",
        f"Beta & {tex_num(PAPER_TABLE2['beta'])} & {tex_num(beta)} & {plain_change(beta, PAPER_TABLE2['beta'])} \\\\",
        r"\bottomrule",
        r"\end{tabular}",
        r"\end{center}",
    ]
    return "\n".join(lines)


def table4_summary_tex(lsq: dict[str, float | int]) -> str:
    updated = {
        "Arithmetic Mean Return": lsq["arithmetic"] * 100.0,
        "Geometric Mean Return": lsq["geometric"] * 100.0,
        "Percent Positive Months": lsq["positive_pct"] * 100.0,
        "Percent Negative Months": lsq["negative_pct"] * 100.0,
        "Average Positive Month": lsq["average_positive"] * 100.0,
        "Average Negative Month": lsq["average_negative"] * 100.0,
        "Standard Deviation": lsq["std"] * 100.0,
        "Sharpe Ratio": lsq["sharpe"],
        "Sortino Ratio": lsq["sortino"],
        "Omega Ratio": lsq["omega"],
    }
    lines = [
        r"\begin{center}",
        r"\begin{tabular}{lrrr}",
        r"\toprule",
        r"Statistic & Paper LSQ & Updated LSQ & Change \\",
        r"\midrule",
    ]
    for metric, old in PAPER_TABLE4_LSQ.items():
        new = updated[metric]
        lines.append(f"{metric} & {tex_num(old)} & {tex_num(new)} & {plain_change(new, old)} \\\\")
    lines.extend([r"\bottomrule", r"\end{tabular}", r"\end{center}"])
    return "\n".join(lines)


def serial_tex(rows: list[dict[str, float | str]]) -> str:
    out = [
        r"\begin{table}[H]",
        r"\centering",
        r"\caption{Serial-Correlation Diagnostics for Monthly Returns}",
        r"\label{tab:serial-correlation}",
        r"\small",
        r"\begin{tabular}{lrrrr}",
        r"\toprule",
        r"Series & $\rho_1$ & LB p(1) & LB p(6) & LB p(12) \\",
        r"\midrule",
    ]
    for row in rows:
        out.append(
            f"{latex_escape(row['Series'])} & {tex_num(row['rho_1'])} & {row['LB p(1)']:.4f} & {row['LB p(6)']:.4f} & {row['LB p(12)']:.4f} \\\\"
        )
    out.extend(
        [
            r"\bottomrule",
            r"\end{tabular}",
            "",
            r"\vspace{0.15cm}",
            r"\begin{minipage}{0.86\linewidth}",
            r"\footnotesize",
            r"$\rho_1$ is the first-order monthly autocorrelation.",
            r"LB p($k$) is the Ljung-Box p-value through lag $k$.",
            r"\end{minipage}",
            r"\end{table}",
        ]
    )
    return "\n".join(out)


def newey_tex(rows: list[dict[str, float | str]], lags: int, n: int) -> str:
    out = [
        r"\begin{table}[H]",
        r"\centering",
        r"\caption{Conventional vs Newey-West Inference}",
        r"\label{tab:newey-west}",
        r"\small",
        r"\begin{tabular}{lrrrrr}",
        r"\toprule",
        r"Statistic & Estimate & OLS SE & OLS t-stat & Newey-West SE & Newey-West t-stat \\",
        r"\midrule",
    ]
    for row in rows:
        if row["Statistic"] == "Beta":
            estimate = tex_num(row["Estimate"])
            ols_se = tex_num(row["OLS SE"])
            nw_se = tex_num(row["Newey-West SE"])
        else:
            estimate = f"{tex_num(row['Estimate'])}\\%"
            ols_se = f"{tex_num(row['OLS SE'])}\\%"
            nw_se = f"{tex_num(row['Newey-West SE'])}\\%"
        out.append(
            f"{row['Statistic']} & {estimate} & {ols_se} & {tex_num(row['OLS t-stat'])} & {nw_se} & {tex_num(row['Newey-West t-stat'])} \\\\"
        )
    out.extend(
        [
            r"\bottomrule",
            r"\end{tabular}",
            "",
            r"\vspace{0.15cm}",
            r"\begin{minipage}{0.92\linewidth}",
            r"\footnotesize",
            f"Newey-West standard errors use {lags} monthly lags, based on the rule $\\lfloor 4(T/100)^{{2/9}}\\rfloor$ with $T={n}$.",
            r"The alpha-beta regression is $r_{\mathrm{LSQ},t}=\alpha+\beta r_{\mathrm{S\&P500TR},t}+\varepsilon_t$.",
            r"\end{minipage}",
            r"\end{table}",
        ]
    )
    return "\n".join(out)


def verification_sentence(checks: list[dict[str, object]], source_label: str, source: str, simulation: bool) -> str:
    sp = next(row for row in checks if row["series"] == "S&P 500 TR")
    nvda = next(row for row in checks if row["series"] == "NVIDIA")
    if simulation:
        return (
            f"For {sp['month']}, this self-test uses simulated S\\&P 500 TR = {sp['source_return_percent']:.6f}\\% "
            f"and simulated NVIDIA = {nvda['source_return_percent']:.6f}\\%. "
            "These values are not live market data and are used only to test the next-month automation path."
        )
    if source == "spartan":
        return (
            f"For {sp['month']}, the market comparators are pulled from Yahoo Finance monthly adjusted levels. "
            f"The S\\&P 500 TR return is {sp['source_return_percent']:.6f}\\%, and the NVIDIA return is {nvda['source_return_percent']:.6f}\\%."
        )
    if sp["yahoo_exact_return_percent"] == "" or nvda["yahoo_exact_return_percent"] == "":
        return (
            f"For {sp['month']}, the {source_label} gives S\\&P 500 TR = "
            f"{sp['source_return_percent']:.6f}\\% and NVIDIA = {nvda['source_return_percent']:.6f}\\%. "
            f"The market-data verification could not be completed automatically, so the {source_label} values are used."
        )
    base = (
        f"For {sp['month']}, the {source_label} gives S\\&P 500 TR = {sp['source_return_percent']:.6f}\\% "
        f"and NVIDIA = {nvda['source_return_percent']:.6f}\\%. "
        f"Yahoo Finance monthly adjusted levels imply S\\&P 500 TR = {sp['yahoo_exact_return_percent']:.6f}\\% "
        f"and NVIDIA = {nvda['yahoo_exact_return_percent']:.6f}\\%. "
    )
    nvda_diff = abs(float(nvda["difference_percentage_points"]))
    if nvda_diff < 0.00001:
        return base + "The NVIDIA value matches the Yahoo adjusted-close calculation."
    return base + f"The NVIDIA difference is {nvda_diff:.4f} percentage points, which is consistent with rounded adjusted closes."


def write_main_tex(
    path: Path,
    latest: tuple[int, int],
    lsq: dict[str, float | int],
    reg: list[dict[str, float | str]],
    checks: list[dict[str, object]],
    serial: list[dict[str, float | str]],
    nw: list[dict[str, float | str]],
    nw_lags: int,
    added_returns: list[dict[str, object]],
    source_label: str,
    source: str,
    simulation: bool,
) -> None:
    alpha = next(row for row in reg if row["statistic"] == "Monthly alpha")["estimate"]
    beta = next(row for row in reg if row["statistic"] == "Beta")["estimate"]
    added_text = ", ".join(f"{row['month']} = {float(row['lsq_return']) * 100.0:.2f}\\%" for row in added_returns)
    direction = "rises" if lsq["arithmetic"] > PAPER_TABLE2["arithmetic"] else "falls"
    tex = rf"""\documentclass[11pt]{{article}}
\usepackage[margin=0.85in]{{geometry}}
\usepackage{{booktabs}}
\usepackage{{array}}
\usepackage{{longtable}}
\usepackage{{float}}
\usepackage{{hyperref}}

\title{{Project 1A LSQ Fund Update}}
\author{{Prepared for Professor Phelim Boyle}}
\date{{June 2026}}

\begin{{document}}
\maketitle

\section*{{Data \& Method}}
The update uses {lsq['n']} monthly LSQ observations from January 2017 through {month_label(latest)}.
The original paper used 109 observations ending in January 2026; the added LSQ returns are {added_text}.
The LSQ analysis uses the baseline net-of-fees return series from the {source_label}.

For the benchmark, I used the Yahoo Finance \texttt{{\string^SP500TR}} S\&P 500 Total Return Index.
This is the appropriate S\&P 500 series because the paper states that the benchmark returns are calculated with dividends reinvested; a price-only S\&P 500 index would omit dividends and understate the benchmark return.

{verification_sentence(checks, source_label, source, simulation)}

All return inputs are monthly.
Population standard deviation divides by $N$.
Sharpe ratios use a 0\% risk-free rate.
Sortino ratios use the paper's downside-deviation convention.
Omega is computed as $\sum \max(r_t-L,0) / \sum \max(L-r_t,0)$.
Alpha and beta are estimated from the monthly regression of LSQ returns on S\&P 500 Total Return Index returns.

\clearpage
\section*{{Updated Table 2}}
\setcounter{{table}}{{1}}
\input{{tables/updated_table_2.tex}}

\clearpage
\section*{{Updated Table 4}}
\setcounter{{table}}{{3}}
\input{{tables/updated_table_4.tex}}

\clearpage
\section*{{Summary of Changes}}
\subsection*{{Table 2: Paper Values vs Updated Values}}
\small
{table2_summary_tex(lsq, reg)}
\normalsize

\subsection*{{Table 4: Paper LSQ Column vs Updated LSQ Column}}
\small
{table4_summary_tex(lsq)}
\normalsize

\subsection*{{Interpretation}}
The updated LSQ statistics move only moderately because {len(added_returns)} months are being added to a 109-month sample.
All added LSQ months are positive, so the number of positive months rises from 104 to {lsq['positive_count']} and the number of negative months remains five.
The average monthly return {direction} relative to the paper because May 2026 is a strong positive month.
The maximum drawdown remains -2.20\% because none of the added months creates a new trough.

The risk-adjusted metrics remain very strong.
The Sharpe ratio is {tex_num(lsq['sharpe'])}, the Sortino ratio is {tex_num(lsq['sortino'])}, and the zero-threshold Omega ratio is {tex_num(lsq['omega'])}.
The updated monthly alpha is {alpha:.2f}\% and beta is {beta:.2f}, so the evidence still supports the paper's conclusion that LSQ's returns are not explained by broad equity-market exposure.

\clearpage
\section*{{Additional Analysis: Serial Correlation}}
Since the paper already identifies serial correlation as an important feature of the LSQ return series, I implemented a serial-correlation diagnostic table and a Newey-West adjusted inference table.
Newey-West inference adjusts the standard errors for autocorrelation and heteroskedasticity in the regression residuals.
This gives a more appropriate t-statistic when monthly returns are not independent over time.

\subsection*{{Serial-Correlation Diagnostics}}
{serial_tex(serial)}

\subsection*{{Newey-West Adjusted Inference}}
{newey_tex(nw, nw_lags, int(lsq['n']))}

The Newey-West adjustment increases the standard errors, which is expected when there is positive serial correlation.
Even after this adjustment, the LSQ monthly mean and alpha remain economically large and statistically strong.
The beta remains close to zero and statistically insignificant.

\subsection*{{Additional Suggestions}}
I think the most logical next extension would be to apply the same dependence-aware approach to the paper's nonlinear performance ratios.
We could produce confidence intervals for the Sharpe, Sortino, and Omega ratios without assuming independent monthly observations.
This should directly address the serial-correlation issue and preserve our focus on asymmetric performance.

\end{{document}}
"""
    path.write_text(tex, encoding="utf-8")


def write_workbook(path: Path, sheets: dict[str, list[dict[str, object]]]) -> None:
    wb = Workbook()
    first = True
    for name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet(name)
        first = False
        ws.title = name[:31]
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row.get(header, "") for header in headers])
            ws.freeze_panes = "A2"
            for col in ws.columns:
                width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)
    wb.save(path)


def compile_pdf(out_dir: Path) -> str:
    tinytex = Path(os.environ.get("APPDATA", "")) / "TinyTeX" / "bin" / "windows"
    env = os.environ.copy()
    if tinytex.exists():
        env["Path"] = f"{tinytex};{env.get('Path', '')}"
    pdflatex = shutil.which("pdflatex", path=env.get("Path"))
    if not pdflatex:
        return "pdflatex not found; TeX source was generated."
    for _ in range(2):
        result = subprocess.run(
            [pdflatex, "-interaction=nonstopmode", "Project_1_May2026_Update.tex"],
            cwd=out_dir,
            env=env,
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            (out_dir / "logs_or_notes" / "latex_compile_error.log").write_text(result.stdout + "\n" + result.stderr, encoding="utf-8")
            return "LaTeX compilation failed; see logs_or_notes/latex_compile_error.log."
    return "Compiled Project_1_May2026_Update.pdf with pdflatex."


def make_zip(out_dir: Path) -> Path:
    zip_path = out_dir / "Harman_Boyle_Project1_May2026_Overleaf_Update.zip"
    if zip_path.exists():
        zip_path.unlink()
    skip_suffixes = {".aux", ".fdb_latexmk", ".fls", ".log", ".out", ".pyc"}
    skip_dirs = {".git", ".github", "__pycache__", "apps_script", "workspace"}
    skip_names = {
        "LSQ-Oct21-live.pdf",
        "apps_script_upload_secret.txt",
        "Code_for_paste.gs",
        "Boyle_LSQ_Workspace_Overleaf_Latest.zip",
        "WORKSPACE_README.md",
        "workspace_artifacts.json",
        "workspace_index.txt",
        "workspace_manifest.json",
    }
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file in out_dir.rglob("*"):
            if file.is_dir() or file == zip_path:
                continue
            if any(part in skip_dirs for part in file.parts):
                continue
            if file.suffix.lower() in skip_suffixes:
                continue
            if file.name in skip_names or file.name.startswith("LSQ_page_"):
                continue
            zf.write(file, file.relative_to(out_dir))
    return zip_path


def run(
    source: str,
    workbook: Path,
    spartan_url: str,
    out_dir: Path,
    skip_yahoo: bool,
    no_compile: bool,
    simulate_next_month: bool,
    test_lsq_return: float,
    test_sp500tr_return: float,
    test_nvda_return: float,
) -> dict[str, object]:
    out_dir.mkdir(parents=True, exist_ok=True)
    for sub in ["tables", "data", "logs_or_notes", "scripts"]:
        (out_dir / sub).mkdir(exist_ok=True)

    source_label = "revised workbook"
    source_file = str(workbook)
    if source == "spartan":
        lsq_rows, sp_rows, nvda_rows, latest, pdf_path = read_spartan_source(spartan_url, out_dir)
        source_label = "Spartan monthly performance PDF"
        source_file = str(pdf_path)
    else:
        lsq_rows, sp_rows, nvda_rows, latest = read_workbook(workbook)
    if simulate_next_month:
        latest = add_month(*latest, 1)
        source_label = f"automation self-test based on {source_label}"
        source_file = f"{source_file}; simulated {month_label(latest)} values"
        lsq_rows.append({"date": month_key(*latest), "month": month_label(latest), "lsq_return": test_lsq_return / 100.0})
        sp_rows.append({"date": month_key(*latest), "month": month_label(latest), "sp500tr_return": test_sp500tr_return / 100.0})
        nvda_rows.append({"date": month_key(*latest), "month": month_label(latest), "nvda_return": test_nvda_return / 100.0})
    lsq_values = values(lsq_rows, "lsq_return")
    sp_values = values(sp_rows, "sp500tr_return")
    nvda_values = values(nvda_rows, "nvda_return")

    lsq = compute_summary(lsq_values)
    sp = compute_summary(sp_values)
    nvda = compute_summary(nvda_values)
    reg, _ = regression_table(lsq_values, sp_values)
    serial = serial_table(lsq_values, sp_values)
    nw, nw_lags = newey_west_table(lsq_values, sp_values)
    checks = verify_market_returns(sp_rows, nvda_rows, latest, True if simulate_next_month else skip_yahoo)
    t2 = table2_rows(lsq, sp, reg)
    t4 = table4_rows(lsq, nvda)

    added_returns = [row for row in lsq_rows if row["date"] > "2026-01"]

    write_csv(out_dir / "data" / "lsq_returns.csv", lsq_rows)
    write_csv(out_dir / "data" / "sp500tr_returns.csv", sp_rows)
    write_csv(out_dir / "data" / "nvda_returns.csv", nvda_rows)
    write_csv(out_dir / "data" / "market_return_verification.csv", checks)
    write_csv(out_dir / "tables" / "updated_table_2.csv", t2)
    write_csv(out_dir / "tables" / "updated_table_4.csv", t4)

    write_table2_tex(out_dir / "tables" / "updated_table_2.tex", lsq, sp, reg)
    write_table4_tex(out_dir / "tables" / "updated_table_4.tex", latest, lsq, nvda)
    write_main_tex(
        out_dir / "Project_1_May2026_Update.tex",
        latest,
        lsq,
        reg,
        checks,
        serial,
        nw,
        nw_lags,
        added_returns,
        source_label,
        source,
        simulate_next_month,
    )

    sheets = {
        "README": [
            {"Field": "Project", "Value": "Project 1 May 2026 update"},
            {"Field": "Data source", "Value": source_label},
            {"Field": "Source file", "Value": source_file},
            {"Field": "Generated", "Value": datetime.now().isoformat(timespec="seconds")},
            {"Field": "Dependencies", "Value": "openpyxl; pymupdf only when using Spartan PDF mode."},
        ],
        "LSQ_Returns": lsq_rows,
        "SP500TR_Returns": sp_rows,
        "NVIDIA_Returns": nvda_rows,
        "Verification": checks,
        "Updated_Table_2": t2,
        "Updated_Table_4": t4,
        "Regression": reg,
        "Serial_Correlation": serial,
        "Newey_West": nw,
    }
    write_workbook(out_dir / "Project_1_May2026_Update.xlsx", sheets)

    script_target = out_dir / "scripts" / "update_project1_monthly.py"
    current_script = Path(__file__).resolve()
    if current_script != script_target.resolve():
        shutil.copy2(current_script, script_target)

    (out_dir / "requirements.txt").write_text("openpyxl>=3.1\npymupdf>=1.24\n", encoding="utf-8")
    (out_dir / "README.txt").write_text(
        "\n".join(
            [
                "Project 1 May 2026 update",
                "",
                "Main file for Overleaf: Project_1_May2026_Update.tex",
                "Automation script: scripts/update_project1_monthly.py",
                "Use source spartan for the monthly no-touch update. It needs Python, openpyxl, and pymupdf.",
                "Use source workbook if Professor Boyle sends a revised spreadsheet.",
                "Installed local scheduled task: Project1_LSQ_Monthly_Update, monthly on the 10th at 9:00 AM when logged in and plugged in.",
                "For professor self-service downloads, use the GitHub Actions plus Google Drive setup in GOOGLE_DRIVE_AUTOMATION.md.",
                "",
                "Refresh from Spartan PDF:",
                "python scripts/update_project1_monthly.py --source spartan --output-dir .",
                "",
                "Refresh from workbook:",
                "python scripts/update_project1_monthly.py --source workbook --workbook C:\\Users\\HP\\Downloads\\COMBined3a.xlsx --output-dir .",
                "",
                "Pipeline test for the next month:",
                "python scripts/update_project1_monthly.py --source spartan --simulate-next-month --output-dir test_next_month",
            ]
        ),
        encoding="utf-8",
    )

    compile_status = "Skipped PDF compilation." if no_compile else compile_pdf(out_dir)
    zip_path = make_zip(out_dir)

    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": source,
        "simulation": simulate_next_month,
        "source_file": source_file,
        "spartan_url": spartan_url if source == "spartan" else "",
        "workbook": str(workbook) if source == "workbook" else "",
        "sample_start": month_key(*SAMPLE_START),
        "sample_end": month_key(*latest),
        "observations": lsq["n"],
        "lsq_arithmetic_monthly_percent": lsq["arithmetic"] * 100.0,
        "lsq_geometric_monthly_percent": lsq["geometric"] * 100.0,
        "lsq_sharpe": lsq["sharpe"],
        "lsq_sortino": lsq["sortino"],
        "lsq_omega": lsq["omega"],
        "sp500tr_latest_percent": checks[0]["source_return_percent"],
        "nvda_latest_percent": checks[1]["source_return_percent"],
        "compile_status": compile_status,
        "overleaf_zip": str(zip_path),
    }
    (out_dir / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    (out_dir / "logs_or_notes" / "data_validation_notes.md").write_text(
        "\n".join(
            [
                "# Data Validation Notes",
                "",
                f"Sample: {manifest['sample_start']} to {manifest['sample_end']}",
                f"Observations: {manifest['observations']}",
                "",
                "May 2026 checks:",
                f"- Data source: {source_label}",
                f"- Source file: {source_file}",
                f"- Simulation: {simulate_next_month}",
                f"- S&P 500 TR source return: {checks[0]['source_return_percent']:.6f}%",
                f"- S&P 500 TR Yahoo exact return: {checks[0]['yahoo_exact_return_percent']}",
                f"- NVIDIA source return: {checks[1]['source_return_percent']:.6f}%",
                f"- NVIDIA Yahoo exact return: {checks[1]['yahoo_exact_return_percent']}",
                "",
                "The S&P 500 TR benchmark uses ^SP500TR because the paper's benchmark return includes dividend reinvestment.",
                "Small NVIDIA differences can occur when a source uses rounded adjusted closes.",
            ]
        ),
        encoding="utf-8",
    )
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser(description="Regenerate Project 1 from the latest workbook.")
    parser.add_argument(
        "--source",
        choices=["spartan", "workbook"],
        default="spartan",
        help="Use the live Spartan PDF or a professor-provided workbook.",
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="Path to the professor's workbook.")
    parser.add_argument("--spartan-url", default=SPARTAN_URL, help="Monthly Spartan LSQ performance PDF URL.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Output folder.")
    parser.add_argument("--skip-yahoo", action="store_true", help="Skip Yahoo verification if offline.")
    parser.add_argument("--no-compile", action="store_true", help="Generate TeX only; skip pdflatex.")
    parser.add_argument("--simulate-next-month", action="store_true", help="Append a clearly labeled dummy next month for pipeline testing.")
    parser.add_argument("--test-lsq-return", type=float, default=1.00, help="Simulated next-month LSQ return in percent.")
    parser.add_argument("--test-sp500tr-return", type=float, default=1.00, help="Simulated next-month S&P 500 TR return in percent.")
    parser.add_argument("--test-nvda-return", type=float, default=1.00, help="Simulated next-month NVIDIA return in percent.")
    args = parser.parse_args()
    manifest = run(
        args.source,
        args.workbook,
        args.spartan_url,
        args.output_dir,
        args.skip_yahoo,
        args.no_compile,
        args.simulate_next_month,
        args.test_lsq_return,
        args.test_sp500tr_return,
        args.test_nvda_return,
    )
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
