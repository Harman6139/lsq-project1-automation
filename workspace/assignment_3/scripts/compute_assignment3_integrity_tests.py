from __future__ import annotations

import csv
import json
import math
import random
import shutil
import subprocess
from collections import Counter
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError as exc:
    raise SystemExit("Install openpyxl first: pip install openpyxl") from exc

try:
    from scipy import stats as scipy_stats
except Exception:
    scipy_stats = None


PROJECT_DIR = Path(__file__).resolve().parents[3]
ASSIGNMENT_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = ASSIGNMENT_DIR / "data"
OUTPUT_TEX = ASSIGNMENT_DIR / "Assignment_3_Statistical_Integrity_Tests.tex"
OUTPUT_PDF = ASSIGNMENT_DIR / "Assignment_3_Statistical_Integrity_Tests.pdf"
OUTPUT_XLSX = ASSIGNMENT_DIR / "Assignment_3_Statistical_Integrity_Tests.xlsx"
OUTPUT_JSON = ASSIGNMENT_DIR / "assignment3_manifest.json"
BOOTSTRAP_REPS = 100_000
BOOTSTRAP_SEED = 20260625


def read_returns(path: Path, key: str) -> list[dict[str, object]]:
    rows = []
    with path.open(newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            rows.append(
                {
                    "date": row["date"],
                    "month": row.get("month", row["date"]),
                    "return_decimal": float(row[key]),
                    "return_percent": float(row[key]) * 100.0,
                }
            )
    return rows


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def normal_cdf(x: float) -> float:
    return 0.5 * (1.0 + math.erf(x / math.sqrt(2.0)))


def normal_two_sided_p(z: float) -> float:
    return math.erfc(abs(z) / math.sqrt(2.0))


def t_two_sided_p(t_stat: float, df: int) -> float:
    if scipy_stats is not None:
        return float(2.0 * scipy_stats.t.sf(abs(t_stat), df))
    return normal_two_sided_p(t_stat)


def binom_cdf(k: int, n: int, p: float) -> float:
    if scipy_stats is not None:
        return float(scipy_stats.binom.cdf(k, n, p))
    if p <= 0:
        return 1.0
    if p >= 1:
        return 1.0 if k >= n else 0.0
    q = 1.0 - p
    prob = q**n
    total = prob
    for i in range(1, k + 1):
        prob *= (n - i + 1) / i * p / q
        total += prob
    return min(max(total, 0.0), 1.0)


def gammaincc(a: float, x: float) -> float:
    if scipy_stats is not None:
        return float(scipy_stats.chi2.sf(2.0 * x, 2.0 * a))
    if x <= 0:
        return 1.0
    eps = 3e-14
    max_iter = 200
    gln = math.lgamma(a)
    if x < a + 1.0:
        ap = a
        delta = 1.0 / a
        total = delta
        for _ in range(max_iter):
            ap += 1.0
            delta *= x / ap
            total += delta
            if abs(delta) < abs(total) * eps:
                p = total * math.exp(-x + a * math.log(x) - gln)
                return max(0.0, min(1.0, 1.0 - p))
        p = total * math.exp(-x + a * math.log(x) - gln)
        return max(0.0, min(1.0, 1.0 - p))

    tiny = 1e-300
    b = x + 1.0 - a
    c = 1.0 / tiny
    d = 1.0 / b if abs(b) > tiny else 1.0 / tiny
    h = d
    for i in range(1, max_iter + 1):
        an = -i * (i - a)
        b += 2.0
        d = an * d + b
        if abs(d) < tiny:
            d = tiny
        c = b + an / c
        if abs(c) < tiny:
            c = tiny
        d = 1.0 / d
        delta = d * c
        h *= delta
        if abs(delta - 1.0) < eps:
            break
    q = math.exp(-x + a * math.log(x) - gln) * h
    return max(0.0, min(1.0, q))


def chi_square_sf(x: float, df: int) -> float:
    if scipy_stats is not None:
        return float(scipy_stats.chi2.sf(x, df))
    return gammaincc(df / 2.0, x / 2.0)


def summary_stats(values: list[float]) -> dict[str, float | int]:
    n = len(values)
    mean = sum(values) / n
    std = math.sqrt(sum((x - mean) ** 2 for x in values) / (n - 1))
    return {
        "n": n,
        "mean": mean,
        "std": std,
        "positive": sum(x > 0.0 for x in values),
        "negative": sum(x < 0.0 for x in values),
        "sharpe": mean / std,
    }


def signs(values: list[float]) -> list[int]:
    return [1 if value > 0 else -1 if value < 0 else 0 for value in values]


def run_count(sign_sequence: list[int]) -> int:
    clean = [x for x in sign_sequence if x != 0]
    if not clean:
        return 0
    return 1 + sum(clean[i] != clean[i - 1] for i in range(1, len(clean)))


def runs_distribution(n_pos: int, n_neg: int) -> dict[int, float]:
    total = math.comb(n_pos + n_neg, n_pos)
    out: dict[int, float] = {}
    max_runs = 2 * min(n_pos, n_neg) + 1
    for r in range(2, max_runs + 1):
        if r % 2 == 0:
            k = r // 2
            count = 2 * math.comb(n_pos - 1, k - 1) * math.comb(n_neg - 1, k - 1)
        else:
            k = (r - 1) // 2
            count = 0
            if n_pos - 1 >= k and n_neg - 1 >= k - 1:
                count += math.comb(n_pos - 1, k) * math.comb(n_neg - 1, k - 1)
            if n_pos - 1 >= k - 1 and n_neg - 1 >= k:
                count += math.comb(n_pos - 1, k - 1) * math.comb(n_neg - 1, k)
        out[r] = count / total
    return out


def runs_test(values: list[float]) -> dict[str, float | int]:
    sign_values = signs(values)
    n_pos = sum(x > 0 for x in sign_values)
    n_neg = sum(x < 0 for x in sign_values)
    n = n_pos + n_neg
    observed = run_count(sign_values)
    expected = 2.0 * n_pos * n_neg / n + 1.0
    variance = 2.0 * n_pos * n_neg * (2.0 * n_pos * n_neg - n) / (n * n * (n - 1.0))
    z = (observed - expected) / math.sqrt(variance)
    distribution = runs_distribution(n_pos, n_neg)
    exact_lower = sum(prob for r, prob in distribution.items() if r <= observed)
    exact_upper = sum(prob for r, prob in distribution.items() if r >= observed)
    exact_two_sided = sum(prob for prob in distribution.values() if prob <= distribution[observed] + 1e-18)
    return {
        "N": n,
        "n_positive": n_pos,
        "n_negative": n_neg,
        "observed_runs": observed,
        "expected_runs": expected,
        "variance": variance,
        "z_stat": z,
        "r_max": 2 * n_neg + 1,
        "exact_lower_tail_p": min(exact_lower, 1.0),
        "exact_upper_tail_p": min(exact_upper, 1.0),
        "exact_two_sided_probability": min(exact_two_sided, 1.0),
        "normal_lower_tail_p": normal_cdf(z),
        "normal_two_sided_p": normal_two_sided_p(z),
    }


def bollen_pool(values: list[float], dates: list[str]) -> dict[str, object]:
    stats = summary_stats(values)
    p_minus = normal_cdf(-float(stats["mean"]) / float(stats["std"]))
    expected_neg = float(stats["n"]) * p_minus
    observed_neg = int(stats["negative"])
    near_pos = sum(0.0 < value < 0.5 for value in values)
    near_neg = sum(-0.5 < value < 0.0 for value in values)
    return {
        "N": stats["n"],
        "mean": stats["mean"],
        "std": stats["std"],
        "p_negative": p_minus,
        "expected_negative": expected_neg,
        "observed_negative": observed_neg,
        "binomial_p_value": binom_cdf(observed_neg, int(stats["n"]), p_minus),
        "near_positive_count": near_pos,
        "near_negative_count": near_neg,
        "near_zero_ratio": math.inf if near_neg == 0 else near_pos / near_neg,
        "negative_months": [
            {"date": date, "return_percent": value}
            for date, value in zip(dates, values)
            if value < 0.0
        ],
    }


def autocorr(values: list[float], lag: int) -> float:
    mean = sum(values) / len(values)
    numerator = sum((values[t] - mean) * (values[t - lag] - mean) for t in range(lag, len(values)))
    denominator = sum((value - mean) ** 2 for value in values)
    return numerator / denominator


def autocorr_table(values: list[float], max_lag: int) -> list[dict[str, float | int]]:
    return [{"lag": lag, "autocorrelation": autocorr(values, lag)} for lag in range(1, max_lag + 1)]


def autocorr_test(values: list[float], lag: int = 1) -> dict[str, float | int]:
    rho = autocorr(values, lag)
    n = len(values)
    t_stat = rho * math.sqrt(n - 2.0) / math.sqrt(1.0 - rho * rho)
    z_stat = rho * math.sqrt(n)
    return {
        "lag": lag,
        "rho": rho,
        "t_stat": t_stat,
        "t_p_value": t_two_sided_p(t_stat, n - 2),
        "z_stat": z_stat,
        "normal_p_value": normal_two_sided_p(z_stat),
    }


def ljung_box(values: list[float], max_lag: int = 12) -> dict[str, float | int]:
    n = len(values)
    q_stat = n * (n + 2.0) * sum((autocorr(values, lag) ** 2) / (n - lag) for lag in range(1, max_lag + 1))
    return {"lag": max_lag, "q_stat": q_stat, "p_value": chi_square_sf(q_stat, max_lag)}


def lo_adjusted_sharpe(values: list[float], max_lag: int = 12) -> dict[str, float | int]:
    stats = summary_stats(values)
    monthly = float(stats["sharpe"])
    naive = math.sqrt(12.0) * monthly
    denominator = 1.0 + 2.0 * sum(autocorr(values, lag) * (1.0 - lag / max_lag) for lag in range(1, max_lag + 1))
    adjusted = naive / math.sqrt(denominator)
    return {
        "lags": max_lag,
        "monthly_sharpe": monthly,
        "naive_annual_sharpe": naive,
        "lo_denominator": denominator,
        "lo_adjusted_annual_sharpe": adjusted,
    }


def bias_ratio(values: list[float], sigma: float | None = None) -> dict[str, float | int]:
    boundary = sigma if sigma is not None else float(summary_stats(values)["std"])
    upper = sum(0.0 <= value <= boundary for value in values)
    lower = sum(-boundary <= value < 0.0 for value in values)
    return {"sigma": boundary, "A": upper, "B": lower, "bias_ratio": upper / (lower + 1.0)}


def expected_bias_ratio_normal(mean: float, sigma: float, n: int) -> float:
    sharpe = mean / sigma
    upper_probability = normal_cdf(1.0 - sharpe) - normal_cdf(-sharpe)
    lower_probability = normal_cdf(-sharpe) - normal_cdf(-1.0 - sharpe)
    return upper_probability / (lower_probability + 1.0 / n)


def bootstrap_bias_ratio(values: list[float], reps: int, seed: int) -> dict[str, float | int]:
    rng = random.Random(seed)
    n = len(values)
    boot = []
    for _ in range(reps):
        sample = [rng.choice(values) for _ in range(n)]
        boot.append(float(bias_ratio(sample)["bias_ratio"]))
    boot.sort()
    observed = float(bias_ratio(values)["bias_ratio"])
    mean = sum(boot) / reps
    median = (boot[reps // 2 - 1] + boot[reps // 2]) / 2.0
    std = math.sqrt(sum((value - mean) ** 2 for value in boot) / reps)
    percentile = sum(value <= observed for value in boot) / reps * 100.0
    return {
        "repetitions": reps,
        "seed": seed,
        "mean": mean,
        "median": median,
        "std": std,
        "p5": boot[int(0.05 * reps)],
        "p95": boot[int(0.95 * reps) - 1],
        "observed_percentile": percentile,
    }


def fmt(value: float, digits: int = 2) -> str:
    if math.isinf(value):
        return "inf"
    return f"{value:.{digits}f}"


def sci(value: float) -> str:
    return f"{value:.2e}"


def tex_escape(text: object) -> str:
    return (
        str(text)
        .replace("\\", r"\textbackslash{}")
        .replace("&", r"\&")
        .replace("%", r"\%")
        .replace("$", r"\$")
        .replace("#", r"\#")
        .replace("_", r"\_")
        .replace("{", r"\{")
        .replace("}", r"\}")
    )


def write_workbook(path: Path, sheets: dict[str, list[dict[str, object]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.properties.creator = "Harman Singh"
    workbook.properties.lastModifiedBy = "Harman Singh"
    first = True
    for sheet_name, rows in sheets.items():
        ws = workbook.active if first else workbook.create_sheet()
        first = False
        ws.title = sheet_name[:31]
        if not rows:
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
        for column in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 10), 55)
    workbook.save(path)


def write_report(path: Path, results: dict[str, object]) -> None:
    summary = results["summary"]
    runs = results["runs"]
    bp = results["bollen_pool"]
    serial = results["serial"]
    bias = results["bias_ratio"]
    neg_months = bp["fund"]["negative_months"]

    negative_text = "; ".join(f"{row['date']} ({row['return_percent']:.2f}\\%)" for row in neg_months)
    tex = rf"""\documentclass[11pt]{{article}}
\usepackage[margin=0.82in]{{geometry}}
\usepackage{{amsmath}}
\usepackage{{booktabs}}
\usepackage{{array}}
\usepackage{{float}}
\newcolumntype{{L}}[1]{{>{{\raggedright\arraybackslash}}p{{#1}}}}

\title{{Project Three: Statistical Integrity Tests for Fund X}}
\author{{Prepared by Harman Singh for Professor Phelim Boyle}}
\date{{June 2026}}

\begin{{document}}
\maketitle

\section*{{Data Box}}
The analysis uses 113 monthly observations from January 2017 through May 2026. Returns are measured in percent per month.

\begin{{table}}[H]
\centering
\caption{{Summary Statistics Used Across Tests}}
\begin{{tabular}}{{lrr}}
\toprule
Statistic & Fund X & S\&P 500 TR \\
\midrule
Sample mean (\%) & {fmt(summary['fund']['mean'])} & {fmt(summary['sp500']['mean'])} \\
Sample standard deviation (\%) & {fmt(summary['fund']['std'])} & {fmt(summary['sp500']['std'])} \\
Positive months & {summary['fund']['positive']} & {summary['sp500']['positive']} \\
Negative months & {summary['fund']['negative']} & {summary['sp500']['negative']} \\
Monthly Sharpe ratio & {fmt(summary['fund']['sharpe'])} & {fmt(summary['sp500']['sharpe'])} \\
\bottomrule
\end{{tabular}}
\end{{table}}

\section*{{Test 1: Wald-Wolfowitz Runs Test}}
Fund X has {runs['observed_runs']} observed sign runs, with {runs['n_positive']} positive months and {runs['n_negative']} negative months.
The expected number of runs under independent signs is {fmt(runs['expected_runs'])}, with variance {fmt(runs['variance'])}.
The maximum possible number of runs is $R_{{max}}=2n_-+1={runs['r_max']}$, and Fund X reaches that maximum.

For the smoothing alternative, the relevant one-sided concern is too few runs.
The exact lower-tail probability $P(R\leq {runs['observed_runs']})$ is {fmt(runs['exact_lower_tail_p'], 3)}.
The normal approximation gives $Z={fmt(runs['z_stat'])}$ and a two-sided p-value of {fmt(runs['normal_two_sided_p'], 3)}.
The test does not flag Fund X. More importantly, the test has little discriminatory power here because only five months are negative; once those five losses are isolated, the run count cannot rise further.

\section*{{Test 2: Bollen-Pool Discontinuity Test}}
Assuming normal returns with Fund X's mean and standard deviation, the probability of a negative month is {fmt(bp['fund']['p_negative'], 3)}.
This implies {fmt(bp['fund']['expected_negative'])} expected negative months, compared with {bp['fund']['observed_negative']} observed negative months.
The binomial lower-tail p-value is {sci(bp['fund']['binomial_p_value'])}.
The S\&P 500 benchmark has $p_-={fmt(bp['sp500']['p_negative'], 3)}$, {fmt(bp['sp500']['expected_negative'])} expected negative months, {bp['sp500']['observed_negative']} observed negative months, and p-value {fmt(bp['sp500']['binomial_p_value'], 3)}.

Fund X's near-zero count is {bp['fund']['near_positive_count']} months in $(0\%,0.5\%)$ and {bp['fund']['near_negative_count']} months in $(-0.5\%,0\%)$, for a near-zero ratio of {fmt(bp['fund']['near_zero_ratio'])}.
The five negative months are {negative_text}.
The normality and independence assumptions both make the formal p-value too severe for Fund X: positive skewness reduces the true left-tail probability, and positive serial correlation lowers the effective sample size.

\section*{{Test 3: Serial Autocorrelation Test}}
The first three raw autocorrelations for Fund X are $\rho_1={fmt(serial['fund_acf'][0]['autocorrelation'], 3)}$, $\rho_2={fmt(serial['fund_acf'][1]['autocorrelation'], 3)}$, and $\rho_3={fmt(serial['fund_acf'][2]['autocorrelation'], 3)}$.
The first-order t-statistic is {fmt(serial['fund_rho1']['t_stat'])}, with p-value {fmt(serial['fund_rho1']['t_p_value'], 4)}.
The S\&P 500 first-order autocorrelation is {fmt(serial['sp500_rho1']['rho'], 3)}.

The active return series, Fund X minus S\&P 500 TR, has autocorrelations $\rho_1={fmt(serial['active_acf'][0]['autocorrelation'], 3)}$, $\rho_2={fmt(serial['active_acf'][1]['autocorrelation'], 3)}$, and $\rho_3={fmt(serial['active_acf'][2]['autocorrelation'], 3)}$.
Its first-order t-statistic is {fmt(serial['active_rho1']['t_stat'])}, with p-value {fmt(serial['active_rho1']['t_p_value'], 3)}.
The Ljung-Box p-value through 12 lags is {fmt(serial['fund_ljung_box']['p_value'], 4)} for raw Fund X returns and {fmt(serial['active_ljung_box']['p_value'], 3)} for active returns.

This pattern supports level-persistence rather than return smoothing. The raw Fund X returns are serially correlated, but the active returns are not.
The naive annualised Sharpe ratio is {fmt(serial['lo_12']['naive_annual_sharpe'])}.
Using Lo's 12-lag adjustment gives {fmt(serial['lo_12']['lo_adjusted_annual_sharpe'])}; using 6 lags gives {fmt(serial['lo_6']['lo_adjusted_annual_sharpe'])}.
Both adjusted values remain economically exceptional.

\section*{{Test 4: Abdulali Bias Ratio}}
Using Fund X's sample standard deviation of {fmt(bias['fund']['sigma'])}\%, the upper interval $[0,\hat\sigma]$ contains {bias['fund']['A']} returns and the lower interval $[-\hat\sigma,0)$ contains {bias['fund']['B']} returns.
The observed bias ratio is therefore {fmt(bias['fund']['bias_ratio'])}.
Under a normal distribution with Fund X's monthly Sharpe ratio, the expected bias ratio is approximately {fmt(bias['expected_normal'])}, already above Abdulali's usual warning thresholds for low-Sharpe hedge funds.

The bootstrap calibration uses {bias['bootstrap']['repetitions']:,} resamples of the 113 observed Fund X returns.
The bootstrap median is {fmt(bias['bootstrap']['median'])}, the mean is {fmt(bias['bootstrap']['mean'])}, and the 90\% interval is [{fmt(bias['bootstrap']['p5'])}, {fmt(bias['bootstrap']['p95'])}].
The observed bias ratio is at the {fmt(bias['bootstrap']['observed_percentile'], 1)}th percentile of the bootstrap distribution.
The S\&P 500 bias ratio is {fmt(bias['sp500']['bias_ratio'])}.
The high Fund X bias ratio is typical for its own return distribution and reflects genuine high positive-return frequency, not a small-positive bunching anomaly.

\clearpage
\section*{{Overall Assessment}}
\begin{{table}}[H]
\centering
\caption{{Summary of Statistical Integrity Tests}}
\begin{{tabular}}{{L{{0.19\linewidth}}L{{0.24\linewidth}}L{{0.25\linewidth}}L{{0.21\linewidth}}}}
\toprule
Test & Key statistic & Assessment & Caveat \\
\midrule
Runs test & $R={runs['observed_runs']}$, $R_{{max}}={runs['r_max']}$ & No evidence of too few runs & Few negative months limits test power \\
Bollen-Pool & $n_-={bp['fund']['observed_negative']}$ vs expected {fmt(bp['fund']['expected_negative'])}; p={sci(bp['fund']['binomial_p_value'])} & Formal test is extreme & Normality and independence assumptions overstate significance \\
Serial autocorrelation & Raw $\rho_1={fmt(serial['fund_rho1']['rho'], 3)}$; active $\rho_1={fmt(serial['active_rho1']['rho'], 3)}$ & Supports level-persistence, not smoothing & Raw returns alone are misleading \\
Bias ratio & BR={fmt(bias['fund']['bias_ratio'])}; bootstrap percentile {fmt(bias['bootstrap']['observed_percentile'], 1)} & Typical for Fund X's own distribution & Abdulali thresholds are not calibrated to high-Sharpe funds \\
\bottomrule
\end{{tabular}}
\end{{table}}

Taken together, the tests do not provide statistical evidence of return smoothing or mechanical avoidance of losses.
They cannot rule out every possible concern about Fund X, but they do not support the specific irregularities these tests were designed to detect.

\end{{document}}
"""
    path.write_text(tex, encoding="utf-8")


def compile_pdf(tex_path: Path) -> str:
    pdflatex = shutil.which("pdflatex")
    if not pdflatex:
        return "pdflatex not found"
    for _ in range(2):
        result = subprocess.run(
            [pdflatex, "-interaction=nonstopmode", tex_path.name],
            cwd=tex_path.parent,
            capture_output=True,
            text=True,
            timeout=120,
        )
        if result.returncode != 0:
            (tex_path.parent / "latex_compile_error.log").write_text(result.stdout + "\n" + result.stderr, encoding="utf-8")
            return "LaTeX compilation failed"
    return "Compiled PDF"


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    lsq_rows = read_returns(PROJECT_DIR / "data" / "lsq_returns.csv", "lsq_return")
    sp_rows = read_returns(PROJECT_DIR / "data" / "sp500tr_returns.csv", "sp500tr_return")
    dates = [str(row["date"]) for row in lsq_rows]
    fund = [float(row["return_percent"]) for row in lsq_rows]
    sp500 = [float(row["return_percent"]) for row in sp_rows]
    active = [a - b for a, b in zip(fund, sp500)]

    write_csv(DATA_DIR / "assignment3_fund_x_returns.csv", lsq_rows)
    write_csv(DATA_DIR / "assignment3_sp500tr_returns.csv", sp_rows)

    summary = {"fund": summary_stats(fund), "sp500": summary_stats(sp500)}
    runs = runs_test(fund)
    bp = {"fund": bollen_pool(fund, dates), "sp500": bollen_pool(sp500, dates)}
    serial = {
        "fund_acf": autocorr_table(fund, 12),
        "active_acf": autocorr_table(active, 12),
        "sp500_acf": autocorr_table(sp500, 12),
        "fund_rho1": autocorr_test(fund, 1),
        "active_rho1": autocorr_test(active, 1),
        "sp500_rho1": autocorr_test(sp500, 1),
        "fund_ljung_box": ljung_box(fund, 12),
        "active_ljung_box": ljung_box(active, 12),
        "lo_12": lo_adjusted_sharpe(fund, 12),
        "lo_6": lo_adjusted_sharpe(fund, 6),
    }
    fund_bias = bias_ratio(fund)
    bias = {
        "fund": fund_bias,
        "expected_normal": expected_bias_ratio_normal(
            float(summary["fund"]["mean"]),
            float(summary["fund"]["std"]),
            int(summary["fund"]["n"]),
        ),
        "bootstrap": bootstrap_bias_ratio(fund, BOOTSTRAP_REPS, BOOTSTRAP_SEED),
        "sp500": bias_ratio(sp500),
    }

    summary_rows = [
        {"Series": "Fund X", **summary["fund"]},
        {"Series": "S&P 500 TR", **summary["sp500"]},
    ]
    runs_rows = [{key: value for key, value in runs.items()}]
    bollen_rows = [
        {key: value for key, value in bp["fund"].items() if key != "negative_months"} | {"Series": "Fund X"},
        {key: value for key, value in bp["sp500"].items() if key != "negative_months"} | {"Series": "S&P 500 TR"},
    ]
    serial_rows = [
        {"Series": "Fund X raw", **serial["fund_rho1"], "ljung_box_q12": serial["fund_ljung_box"]["q_stat"], "ljung_box_p12": serial["fund_ljung_box"]["p_value"]},
        {"Series": "Active return", **serial["active_rho1"], "ljung_box_q12": serial["active_ljung_box"]["q_stat"], "ljung_box_p12": serial["active_ljung_box"]["p_value"]},
        {"Series": "S&P 500 TR", **serial["sp500_rho1"], "ljung_box_q12": "", "ljung_box_p12": ""},
    ]
    acf_rows = []
    for lag in range(1, 13):
        acf_rows.append(
            {
                "lag": lag,
                "Fund X raw": serial["fund_acf"][lag - 1]["autocorrelation"],
                "Active return": serial["active_acf"][lag - 1]["autocorrelation"],
                "S&P 500 TR": serial["sp500_acf"][lag - 1]["autocorrelation"],
            }
        )
    bias_rows = [
        {"Series": "Fund X", **bias["fund"], "expected_normal_br": bias["expected_normal"]},
        {"Series": "S&P 500 TR", **bias["sp500"], "expected_normal_br": ""},
    ]
    bootstrap_rows = [{key: value for key, value in bias["bootstrap"].items()}]
    negative_rows = bp["fund"]["negative_months"]
    assessment_rows = [
        {"Test": "Runs test", "Key statistic": f"R={runs['observed_runs']}, Rmax={runs['r_max']}", "Assessment": "No evidence of too few runs", "Caveat": "Few negative months limits test power"},
        {"Test": "Bollen-Pool", "Key statistic": f"p={bp['fund']['binomial_p_value']:.2e}", "Assessment": "Formal result is extreme", "Caveat": "Assumptions overstate significance"},
        {"Test": "Serial autocorrelation", "Key statistic": f"raw rho1={serial['fund_rho1']['rho']:.3f}; active rho1={serial['active_rho1']['rho']:.3f}", "Assessment": "Level-persistence, not smoothing", "Caveat": "Raw returns alone are misleading"},
        {"Test": "Bias ratio", "Key statistic": f"BR={bias['fund']['bias_ratio']:.2f}; percentile={bias['bootstrap']['observed_percentile']:.1f}", "Assessment": "Typical for Fund X distribution", "Caveat": "Abdulali thresholds do not fit high-Sharpe funds"},
    ]

    write_workbook(
        OUTPUT_XLSX,
        {
            "Data Box": summary_rows,
            "Runs Test": runs_rows,
            "Bollen Pool": bollen_rows,
            "Negative Months": negative_rows,
            "Serial Tests": serial_rows,
            "Autocorrelations": acf_rows,
            "Lo Sharpe": [serial["lo_12"], serial["lo_6"]],
            "Bias Ratio": bias_rows,
            "Bias Bootstrap": bootstrap_rows,
            "Assessment": assessment_rows,
        },
    )

    results = {"summary": summary, "runs": runs, "bollen_pool": bp, "serial": serial, "bias_ratio": bias}
    write_report(OUTPUT_TEX, results)
    compile_status = compile_pdf(OUTPUT_TEX)

    manifest = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "sample": "January 2017 to May 2026",
        "observations": len(fund),
        "bootstrap_reps": BOOTSTRAP_REPS,
        "bootstrap_seed": BOOTSTRAP_SEED,
        "compile_status": compile_status,
        "outputs": {
            "pdf": str(OUTPUT_PDF),
            "tex": str(OUTPUT_TEX),
            "workbook": str(OUTPUT_XLSX),
            "script": str(Path(__file__).resolve()),
        },
        "headline_results": {
            "fund_mean_percent": summary["fund"]["mean"],
            "fund_std_percent": summary["fund"]["std"],
            "runs_observed": runs["observed_runs"],
            "bollen_pool_p_value": bp["fund"]["binomial_p_value"],
            "active_rho1": serial["active_rho1"]["rho"],
            "bias_ratio": bias["fund"]["bias_ratio"],
            "bias_bootstrap_percentile": bias["bootstrap"]["observed_percentile"],
        },
    }
    OUTPUT_JSON.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(manifest, indent=2))


if __name__ == "__main__":
    main()
